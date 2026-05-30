import pytest
import openpyxl
from main import load_recipients, load_config


def test_load_recipients(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["company", "full_name", "email"])
    ws.append(["Acme", "Jane Smith", "jane@acme.com"])
    ws.append(["Beta Corp", "John Doe", "john@beta.com"])
    ws.append([None, None, None])  # empty row — must be skipped
    path = str(tmp_path / "recipients.xlsx")
    wb.save(path)

    recipients = load_recipients(path)
    assert len(recipients) == 2
    assert recipients[0]["full_name"] == "Jane Smith"
    assert recipients[0]["company"] == "Acme"
    assert recipients[1]["email"] == "john@beta.com"


def test_load_recipients_normalizes_headers(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Company", "Full Name", "Email Id"])  # real-world header casing
    ws.append(["Acme", "Jane Smith", "jane@acme.com"])
    path = str(tmp_path / "recipients.xlsx")
    wb.save(path)

    recipients = load_recipients(path)
    assert recipients[0]["company"] == "Acme"
    assert recipients[0]["full_name"] == "Jane Smith"
    assert recipients[0]["email"] == "jane@acme.com"


def test_load_config(tmp_path):
    config_file = tmp_path / "config.yml"
    config_file.write_text(
        "resume_path: attachments/resume.pdf\ncookies_path: cookies.json\n",
        encoding="utf-8",
    )
    config = load_config(str(config_file))
    assert config["resume_path"] == "attachments/resume.pdf"
    assert config["cookies_path"] == "cookies.json"


def test_load_recipients_skips_rows_without_email(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["company", "full_name", "email"])
    ws.append(["Acme", "Jane Smith", "jane@acme.com"])
    ws.append(["Beta", "John Doe", ""])          # empty email — skip
    ws.append(["Gamma", "Alice Brown", None])    # None email — skip
    path = str(tmp_path / "recipients.xlsx")
    wb.save(path)

    recipients = load_recipients(path)
    assert len(recipients) == 1
    assert recipients[0]["email"] == "jane@acme.com"
