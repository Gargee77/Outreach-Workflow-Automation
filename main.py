import argparse
import json
from pathlib import Path

import openpyxl
import yaml

from src.template import render_email
from src.gmail_api_creator import GmailApiDraftCreator

# Maps raw Excel header variants → normalized key used in code
_HEADER_ALIASES = {
    "full name": "full_name",
    "email id": "email",
    "email_id": "email",
}


def load_config(config_path: str = "config.yml") -> dict:
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_recipients(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    raw_headers = [str(cell.value or "").lower().strip() for cell in ws[1]]
    headers = [_HEADER_ALIASES.get(h, h.replace(" ", "_")) for h in raw_headers]
    recipients = []
    seen_emails = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        email = str(row_dict.get("email") or "").strip().lower()
        if email and email not in seen_emails:
            seen_emails.add(email)
            recipients.append(row_dict)
    return recipients


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="recipients.xlsx", help="Recipients Excel file")
    parser.add_argument("--skip", type=int, default=0, help="Skip first N recipients")
    args = parser.parse_args()

    base_dir = Path(__file__).parent
    config = load_config(str(base_dir / "config.yml"))
    recipients = load_recipients(str(base_dir / args.file))

    if not recipients:
        print(f"No recipients found in {args.file}. Exiting.")
        return

    template_path = str(base_dir / "email_body.md")
    companies_dir = str(base_dir / "companies")
    attachment_path = str(base_dir / config["resume_path"])

    recipients = recipients[args.skip:]
    print(f"Loaded {len(recipients)} recipients (skipping first {args.skip})")

    draft_ids_path = base_dir / "draft_ids.json"
    existing_ids = json.loads(draft_ids_path.read_text()) if draft_ids_path.exists() else []

    new_ids = []
    with GmailApiDraftCreator(credentials_path=str(base_dir / "credentials.json")) as creator:
        for i, recipient in enumerate(recipients, args.skip + 1):
            company = str(recipient["company"]).strip()
            full_name = str(recipient["full_name"])
            email = str(recipient["email"])

            print(f"[{i}/{len(recipients)}] {full_name} <{email}> ({company}) ...", end=" ", flush=True)
            try:
                rendered = render_email(full_name, company, template_path, companies_dir)
                draft_id = creator.create_draft(
                    to=email,
                    subject=rendered["subject"],
                    html_body=rendered["html_body"],
                    attachment_path=attachment_path,
                )
                new_ids.append(draft_id)
                print("draft saved")
            except Exception as e:
                print(f"SKIPPED ({e})")

    all_ids = existing_ids + new_ids
    draft_ids_path.write_text(json.dumps(all_ids, indent=2))
    print(f"\nDone. {len(new_ids)} drafts saved. Run 'python send_drafts.py' when ready to send.")


if __name__ == "__main__":
    main()
