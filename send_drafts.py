import json
from pathlib import Path

import openpyxl

from src.gmail_api_creator import GmailApiDraftCreator

DRAFT_IDS_PATH = Path(__file__).parent / "draft_ids.json"
_HEADER_ALIASES = {"full name": "full_name", "email id": "email", "email_id": "email"}


def load_recipient_emails(xlsx_path: str) -> set:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    raw_headers = [str(cell.value or "").lower().strip() for cell in ws[1]]
    headers = [_HEADER_ALIASES.get(h, h.replace(" ", "_")) for h in raw_headers]
    emails = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get("email"):
            emails.add(str(row_dict["email"]).strip().lower())
    return emails


def get_cold_email_draft_ids(creator, recipient_emails: set) -> list:
    draft_ids = []
    page_token = None
    while True:
        params = {"userId": "me", "maxResults": 500}
        if page_token:
            params["pageToken"] = page_token
        result = creator._service.users().drafts().list(**params).execute()
        for d in result.get("drafts", []):
            meta = creator._service.users().drafts().get(
                userId="me", id=d["id"], format="metadata"
            ).execute()
            to_header = next(
                (h["value"] for h in meta["message"].get("payload", {}).get("headers", []) if h["name"] == "To"),
                ""
            )
            if to_header.strip().lower() in recipient_emails:
                draft_ids.append(d["id"])
        page_token = result.get("nextPageToken")
        if not page_token:
            break
    return draft_ids


def main():
    base_dir = Path(__file__).parent

    with GmailApiDraftCreator(credentials_path=str(base_dir / "credentials.json")) as creator:
        if DRAFT_IDS_PATH.exists() and json.loads(DRAFT_IDS_PATH.read_text()):
            draft_ids = json.loads(DRAFT_IDS_PATH.read_text())
            print(f"Found {len(draft_ids)} tracked drafts.")
        else:
            print("Scanning Gmail drafts for cold emails (matching recipients.xlsx)...")
            recipient_emails = load_recipient_emails(str(base_dir / "recipients.xlsx"))
            draft_ids = get_cold_email_draft_ids(creator, recipient_emails)
            print(f"Found {len(draft_ids)} matching drafts.")

        if not draft_ids:
            print("Nothing to send.")
            return

        confirm = input(f"Type 'send' to confirm sending {len(draft_ids)} cold email drafts: ").strip().lower()
        if confirm != "send":
            print("Cancelled.")
            return

        sent, failed = 0, []
        for i, draft_id in enumerate(draft_ids, 1):
            print(f"[{i}/{len(draft_ids)}] Sending ...", end=" ", flush=True)
            try:
                creator.send_draft(draft_id)
                print("sent")
                sent += 1
            except Exception as e:
                print(f"FAILED ({e})")
                failed.append(draft_id)

    DRAFT_IDS_PATH.write_text(json.dumps(failed, indent=2))
    print(f"\nDone. {sent} sent, {len(failed)} failed.")
    if failed:
        print(f"{len(failed)} failed IDs kept in draft_ids.json for retry.")


if __name__ == "__main__":
    main()
